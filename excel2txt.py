#!/usr/bin/env python3
"""
Author : Ken Youens-Clark <kyclark@gmail.com>
Date   : 2019-09-26
Purpose: Convert Excel files to delimited text
"""

import argparse
import csv
import os
import re
import sys
from openpyxl import load_workbook
from typing import Any, Optional, NamedTuple, TextIO, List

VERSION = '0.2.1'


class Args(NamedTuple):
    file: List[TextIO]
    out_dir: str
    delimiter: str
    make_dirs: bool
    normalize_headers: bool


# --------------------------------------------------
def get_args() -> Args:
    """ Get command-line arguments """

    parser = argparse.ArgumentParser(
        prog='excel2txt',
        description='Convert Excel files to delimited text',
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)

    parser.add_argument('file',
                        metavar='FILE',
                        type=argparse.FileType('r'),
                        nargs='+',
                        help='Input Excel file(s)')

    parser.add_argument('-o',
                        '--outdir',
                        help='Output directory',
                        metavar='str',
                        type=str,
                        default=os.getcwd())

    parser.add_argument('-d',
                        '--delimiter',
                        help='Delimiter for output file',
                        metavar='str',
                        type=str,
                        default='\t')

    parser.add_argument('-D',
                        '--mkdirs',
                        help='Create separate directories for output files',
                        action='store_true')

    parser.add_argument('-n',
                        '--normalize',
                        help='Normalize headers',
                        action='store_true')

    parser.add_argument('--version',
                        action='version',
                        version=f'%(prog)s {VERSION}')

    args = parser.parse_args()

    if not os.path.isabs(args.outdir):
        args.outdir = os.path.abspath(args.outdir)

    if not os.path.isdir(args.outdir):
        os.makedirs(args.outdir)

    return Args(file=args.file,
                out_dir=args.outdir,
                delimiter=args.delimiter,
                make_dirs=args.mkdirs,
                normalize_headers=args.normalize)


# --------------------------------------------------
def main() -> None:
    """ Make a jazz noise here """

    args = get_args()

    for i, fh in enumerate(args.file, start=1):
        print(f'{i:3}: {fh.name}')
        if not process(fh, args):
            print(f'Something amiss with {fh.name}', file=sys.stderr)

    print(f'Done, see output in "{args.out_dir}".')


# --------------------------------------------------
def process(fh: TextIO, args: Args) -> bool:
    """ Process a file """

    file = fh.name
    fh.close()
    basename, _ = os.path.splitext(os.path.basename(file))
    basename = normalize(basename)
    wb = load_workbook(file)

    for ws in wb:
        ws_name = normalize(ws.title)
        if not ws_name:
            continue

        out_ext = '.csv' if args.delimiter == ',' else '.txt'
        out_file = '__'.join([basename, ws_name]) + out_ext
        out_dir = os.path.join(args.out_dir,
                               basename) if args.make_dirs else args.out_dir

        if not os.path.isdir(out_dir):
            os.makedirs(out_dir)

        out_path = os.path.join(out_dir, out_file)

        with open(out_path, 'wt') as out_fh:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Clean up headers, get rid of bogus/empty columns
            fieldnames = list(rows.pop(0))
            if args.normalize_headers:
                fieldnames = list(map(normalize, fieldnames))
            while fieldnames and fieldnames[-1] is None:
                fieldnames.pop()

            writer = csv.DictWriter(out_fh,
                                    fieldnames=fieldnames,
                                    delimiter=args.delimiter)
            writer.writeheader()

            for row in rows:
                # Only use defined columns
                data = list(map(cell_norm, row[:len(fieldnames)]))

                # Skip empty rows
                if any(filter(None, data)):
                    writer.writerow(dict(zip(fieldnames, data)))

        # Remove empty worksheets
        if os.path.getsize(out_path) == 0:
            os.remove(out_path)

    return True


# --------------------------------------------------
def normalize(s: Optional[str]) -> str:
    """
    Remove funky bits from strings, change CamelCase to snake_case
    """

    if s:
        while True:
            match = re.search('(.*)([a-z])([A-Z].*)', s)
            if match:
                s = match.group(1) + match.group(2) + '_' + match.group(3)
            else:
                break

        s = re.sub(r'\s+', '_', s.lower())
        s = re.sub(r'[^a-z0-9_]', '', s)
        return re.sub(r'[_]+', '_', s)
    else:
        return ''


# --------------------------------------------------
def cell_norm(val: Any) -> str:
    """ Normalize None/etc """

    return '' if val is None else str(val)


# --------------------------------------------------
if __name__ == '__main__':
    main()
